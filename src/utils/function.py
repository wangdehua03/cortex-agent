import os
import re
import shlex
import subprocess
import sys
import threading
import time
from pathlib import Path
from typing import Union, List, Optional, Callable
import openpyxl, xlrd, docx
from openpyxl.styles import Alignment
from pypdf import PdfReader

sys.path.append(os.path.join(os.path.dirname(__file__), '../'))
from config.config import WORKDIR
from src.utils.shell_backend import get_backend, CommandParseResult


# 全局用户输入队列注册，供 _ask_permission 使用
_user_input_queue = None


def _get_user_input_queue():
    if _user_input_queue is None:
        raise LookupError("UserInputQueue not registered")
    return _user_input_queue


def _set_user_input_queue(uiq):
    global _user_input_queue
    _user_input_queue = uiq







def _validate_command(
    command: str,
    *,
    on_dangerous_command: Optional[Callable[[str, str], bool]] = None,
) -> Optional[str]:
    """
    校验用户命令是否安全。不安全则返回错误消息字符串，安全返回 None。

    策略：
    1. 敏感命令：不在白名单中的命令，向用户申请权限（若提供回调）
    2. 危险命令：sudo/shutdown/mkfs 等，向用户申请权限并明确警告（若提供回调）
    3. 白名单命令：直接放行
    4. 高风险命令参数校验：rm、chmod、chown、dd 等操作路径必须在工作目录内

    命令解析已委托给当前平台的 shell 后端（见 shell_backend.py::ShellBackend.parse_command）。
    命令分级清单同样由后端提供。
    """
    backend = get_backend()
    dangerous_cmds = backend.dangerous_cmds
    sensitive_cmds = backend.sensitive_cmds
    safe_cmds = backend.safe_cmds

    result = backend.parse_command(command)
    all_cmd_bases = result.command_names
    seg_tokens = result.first_segment_tokens

    if not all_cmd_bases:
        return "Error: Empty command"

    cmd_base = backend.normalize_cmd(all_cmd_bases[0])

    # Collect all rejection reasons first, then ask user once (not per-command)
    # Group by category: dangerous, sensitive, not in whitelist
    dangerous_found = []
    sensitive_found = []
    unwhitelisted_found = []
    for cb in all_cmd_bases:
        if not cb:
            continue
        cb = backend.normalize_cmd(cb)
        if cb in dangerous_cmds:
            dangerous_found.append(cb)
        elif cb in sensitive_cmds:
            sensitive_found.append(cb)
        elif cb not in safe_cmds:
            unwhitelisted_found.append(cb)

    # Collect all rejection reasons: command-level + path-level, then ask once.
    # Command-level reasons
    command_reason_parts = []
    if dangerous_found:
        command_reason_parts.append(f"Dangerous command: {';'.join(dict.fromkeys(dangerous_found))}")
    if sensitive_found:
        command_reason_parts.append(f"Sensitive command: {';'.join(dict.fromkeys(sensitive_found))}")
    if unwhitelisted_found:
        command_reason_parts.append(f"Command not in whitelist: {';'.join(dict.fromkeys(unwhitelisted_found))}")

    # Path-level reasons (extract once, check all risky paths)
    path_reason_parts = []
    risky_cmds = backend.risky_cmds
    if cmd_base in risky_cmds:
        # 绝对路径判定：Unix 根路径、相对上级、home、Windows 盘符、UNC
        path_re = re.compile(r'^(/|\.{1,3}/|~|[a-zA-Z]:[\\/]|\\\\)')
        for token in seg_tokens[1:]:
            if token.startswith("-"):
                continue
            expanded = os.path.expanduser(token)
            if path_re.match(token) or os.path.isabs(expanded):
                resolved = os.path.realpath(expanded)
                if not _is_within_workspace(resolved):
                    path_reason_parts.append(f"Path outside workspace: {token}")

    # Deduplicate path reasons
    path_reason_parts = list(dict.fromkeys(path_reason_parts))

    all_reasons = command_reason_parts + path_reason_parts
    if all_reasons:
        combined_reason = "\n".join(f"- {r}" for r in all_reasons)
        if on_dangerous_command and on_dangerous_command(combined_reason, command):
            pass  # user approved
        else:
            return f"Error: {combined_reason} — permission denied"

    return None

def _ask_permission(reason: str, command: str) -> bool:
    """默认的交互式权限申请回调：打印警告，向用户请求确认。

    尝试通过全局注册的用户输入队列获取确认（队列化模式）；
    未注册时回退到直接 input()。
    """
    try:
        _uiq = _get_user_input_queue()
    except LookupError:
        _uiq = None

    for line in reason.split("\n"):
        print(f"\033[33m⚠  {line}\033[0m")
    print(f"   命令: {command}")
    if _uiq is not None:
        ans = _uiq.prompt_and_wait("   允许执行吗? (y/N): ")
        return ans is not None and ans.strip().lower() in ("y", "yes")
    else:
        ans = input("   允许执行吗? (y/N): ")
        return ans.strip().lower() in ("y", "yes")


def _ask_path_permission(path: str, reason: str) -> bool:
    """文件路径超出工作区时的权限申请回调。

    尝试通过全局注册的用户输入队列获取确认（队列化模式）；
    未注册时回退到直接 input()。
    """
    try:
        _uiq = _get_user_input_queue()
    except LookupError:
        _uiq = None

    print(f"\033[33m⚠  {reason}\033[0m")
    print(f"   路径: {path}")
    if _uiq is not None:
        ans = _uiq.prompt_and_wait("   允许访问吗? (y/N): ")
        return ans is not None and ans.strip().lower() in ("y", "yes")
    else:
        ans = input("   允许访问吗? (y/N): ")
        return ans.strip().lower() in ("y", "yes")


def _is_interactive_command(command: str) -> bool:
    """判断命令是否需要交互式输入（如密码、确认等）。清单由平台后端提供。"""
    backend = get_backend()
    try:
        tokens = shlex.split(command.strip().splitlines()[0])
        if tokens:
            cmd = backend.normalize_cmd(tokens[0])
            if cmd in backend.interactive_cmds:
                return True
            # Check for compound commands like 'docker login'
            cmd_full = ' '.join(tokens[:2]) if len(tokens) > 1 else cmd
            if cmd_full in backend.interactive_cmds:
                return True
    except (ValueError, IndexError):
        pass

    # 也检查命令中是否包含常见交互式参数模式
    if re.search(r'(ssh|scp|sftp|ftp|telnet)\s+\S+@', command):
        return True
    return False


def run_shell_command(command: str, permission_callback: Optional[Callable[[str, str], bool]] = None) -> str:
    """Execute a shell command via the platform backend (bash on Linux/macOS,
    PowerShell on Windows).  May raise subprocess.TimeoutExpired if the command
    takes longer than 120s — the caller (_process_single_tool) is responsible
    for deciding whether to promote to background."""
    callback = permission_callback if permission_callback is not None else _ask_permission
    err = _validate_command(command, on_dangerous_command=callback)
    if err:
        return err

    backend = get_backend()
    if _is_interactive_command(command):
        return backend.run_interactive(command, str(WORKDIR), _get_user_input_queue)

    try:
        r = backend.run(command, str(WORKDIR), timeout=120)
    except (FileNotFoundError, OSError) as e:
        return f"Error: {e}"
    out = ((r.stdout or "") + (r.stderr or "")).strip()
    suffix = ""
    if r.returncode != 0:
        suffix = f"\n(return code: {r.returncode})"
    return (out[:50000] if out else "(no output)") + suffix


def _is_within_workspace(resolved: str) -> bool:
    """判断 resolved 绝对路径是否位于 WORKDIR 内（Windows 下大小写不敏感比较）。"""
    root = os.path.normcase(os.path.realpath(str(WORKDIR)))
    target = os.path.normcase(resolved)
    return target == root or target.startswith(root + os.sep)



def safe_path(p: Union[str, Path]) -> Path:
    """防止操作目录逃逸（Windows 下大小写不敏感比较）。

    如果路径超出工作区，先向用户申请权限；用户拒绝或交互失败时抛出 ValueError。
    """
    root = WORKDIR.resolve()
    path = (root / Path(p)).resolve()
    root_s = os.path.normcase(str(root))
    path_s = os.path.normcase(str(path))
    if path_s != root_s and not path_s.startswith(root_s + os.sep):
        reason = f"Path outside workspace: {p}"
        if not _ask_path_permission(p, reason):
            raise ValueError(f"Path escapes workspace: {p}")
    return path

def _decode_text(data: bytes) -> tuple[str, str | None]:
    """
    尝试用常见编码解码字节流。
    返回 (text, detected_encoding)。
    若全部失败，使用 latin-1 兜底（保证不抛异常）。
    """
    # BOM 优先：带 BOM 的文件几乎可确定编码
    if data.startswith(b"\xff\xfe\x00\x00"):
        return data.decode("utf-32-le"), "utf-32-le"
    if data.startswith(b"\x00\x00\xfe\xff"):
        return data.decode("utf-32-be"), "utf-32-be"
    if data.startswith(b"\xff\xfe"):
        return data.decode("utf-16-le"), "utf-16-le"
    if data.startswith(b"\xfe\xff"):
        return data.decode("utf-16-be"), "utf-16-be"

    candidates = ["utf-8", "gbk", "gb2312"]
    for enc in candidates:
        try:
            return data.decode(enc), enc
        except (UnicodeDecodeError, LookupError):
            continue
    # 最终兜底：latin-1 每个字节都有对应字符，不会抛异常
    return data.decode("latin-1"), "latin-1"


def run_read(path: str, limit: int = None) -> str:
    try:
        fp = safe_path(path)
        text, detected = _decode_text(fp.read_bytes())
        lines = text.splitlines()
        if limit and limit < len(lines):
            lines = lines[:limit] + [f"...({len(lines) - limit} more lines)"]
        result = "\n".join(lines)[:50000]
        if detected and detected != "utf-8":
            result = f"[detected encoding: {detected}]\n{result}"
        return result
    except Exception as e:
        return f"Error: {e}"
    
def run_write(path: str, content: str) -> str:
    try:
        fp = safe_path(path)
        fp.parent.mkdir(parents=True, exist_ok=True)
        fp.write_text(content)
        return f"Wrote {len(content)} bytes to {path}"
    except Exception as e:
        return f"Error: {e}"
    

def run_edit(path: str, old_text: str, new_text: str) -> str:
    try:
        fp = safe_path(path)
        content = fp.read_text()
        if old_text not in content:
            return f"Error: Text not found in {path}"
        fp.write_text(content.replace(old_text, new_text, 1))
        return f"Edited {path}"
    except Exception as e:
        return f"Error: {e}"


def _escape_md_cell(value: str) -> str:
    """
    处理excel单元格中的特殊字符：
    1、换行符 -> <br> 标签
    2、管道符 | -> 转义为 \|
    3、首尾空格保留（用HTML实体或不换行空格）
    """
    if value is None:
        return ""
    
    # 转换为字符串
    text = str(value)
    text = text.replace("|","\\|")
    text = text.replace("\n","<br>")
    text = text.replace("\r\n","<br>")
    text = text.replace("\r","<br>")

    return text

    
def read_excel_to_md(file_path: Union[str, Path]) -> str:
    """
    读取 xls 或 xlsx 文件， 返回 Markdown 表格格式
    Args:
        file_path (Union[str, Path]): Excel 文件路径

    Returns:
        str: Markdown 格式的表格字符串
    """
    file_path = safe_path(str(file_path))
    if not file_path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")
    
    suffix = file_path.suffix.lower()

    if suffix == '.xlsx':
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        rows = []
        max_col = sheet.max_column
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row,
                                   min_col=1, max_col=max_col,
                                   values_only=False):
            row_data = []
            for cell in row:
                # 获取单元格值， 处理合并单元格
                value = cell.value if cell.value is not None else ""
                row_data.append(value)
            rows.append(row_data)

    elif suffix == '.xls':
        wb = xlrd.open_workbook(file_path, formatting_info=True)
        sheet = wb.sheet_by_index(0)

        rows = []
        for row_idx in range(sheet.nrows):
            row = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                value = cell.value if cell.value is not None else ""
                row.append(value)
            rows.append(row)
    else:
        raise ValueError(f"不支持的文件格式：{suffix}, 仅支持 .xls 和 .xlsx")
    
    if not rows:
        return ""
    
    # 转换为md表格
    md_lines = []
    # 表头
    header = [_escape_md_cell(cell) for cell in rows[0]]
    md_lines.append("|" + "|".join(header) + "|")
    # 分隔符
    md_lines.append("|"+"|".join(["---"] * len(header)) + "|")
    # 数据行
    for row in rows[1:]:
        # 补齐列数（防止某些行列数不足）
        while len(row) < len(header):
            row.append("")

        cells = [_escape_md_cell(cell) for cell in row[:len(header)]]
        md_lines.append("|" + "|".join(cells) + "|")

    return "\n".join(md_lines)


def read_docx_to_text(path: Union[str, Path]) -> str:
    """
    读取 Word 文档 (.docx)，返回纯文本内容。
    会保留段落结构，并将表格转换为 '|' 分隔的文本行。
    """
    file_path = safe_path(path)
    document = docx.Document(file_path)
    lines: List[str] = []

    for para in document.paragraphs:
        text = para.text.strip()
        if text:
            lines.append(text)

    for table in document.tables:
        lines.append("\n[Table]")
        for row in table.rows:
            cells = [cell.text.replace("\n", " ").replace("|", "\\|") for cell in row.cells]
            lines.append(" | ".join(cells))
        lines.append("[End Table]\n")

    result = "\n".join(lines)
    return result[:50000]


def read_pdf_to_text(path: Union[str, Path], max_pages: int = None) -> str:
    """
    读取 PDF 文件，返回提取的文本内容。
    默认读取所有页面；可通过 max_pages 限制页数。
    """
    file_path = safe_path(path)
    reader = PdfReader(str(file_path))
    total_pages = len(reader.pages)
    end_page = min(max_pages, total_pages) if max_pages else total_pages

    lines: List[str] = []
    for i in range(end_page):
        page = reader.pages[i]
        text = page.extract_text() or ""
        lines.append(f"--- Page {i + 1} ---")
        if text.strip():
            lines.append(text.strip())

    result = "\n".join(lines)
    if max_pages and end_page < total_pages:
        result += f"\n\n...(truncated after {max_pages} of {total_pages} pages)"
    return result[:50000]


def _parse_md_table(md_content: str) -> List[List[str]]:
    """
    解析markdown表格，处理<br>标签还原为换行符
    """
    lines = [line.rstrip() for line in md_content.strip().split('\n') if line.strip()]
    rows = []
    for line in lines:
        # 跳过分隔行
        # 检查是否全是分隔符和空格
        stripped = line.replace("|","").replace("-","").replace(" ","").replace(":","")
        if not stripped:
            continue

        # 分割单元格
        # 处理转义的管道符 \|, 先替换为临时标记
        temp_line = line.replace('\\|','<<PIPE>>')
        # 按 | 分割，保留空单元格（仅去掉因首尾 | 产生的空字符串）
        parts = temp_line.split('|')
        if temp_line.startswith('|'):
            parts = parts[1:]
        if temp_line.endswith('|'):
            parts = parts[:-1]
        cells = [cell.strip() for cell in parts]
        # 还原管道符和换行
        cells = [cell.replace('<<PIPE>>','|').replace('<br>','\n') for cell in cells]

        if cells:
            rows.append(cells)

    return rows

def write_md_to_excel(file_path: Union[str, Path], md_content: str) -> None:
    """
    将 Markdown 表格写入xlsx文件
    支持<br>标签还原为单元格内换行
    """
    file_path = safe_path(str(file_path))
    suffix = file_path.suffix.lower()
    rows = _parse_md_table(md_content)

    if not rows:
        raise ValueError("Markdown 内容为空或格式不正确")
    
    if suffix == '.xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active

        for row_idx, row in enumerate(rows, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                # 如果内容包含换行， 设置自动换行对齐
                if isinstance(cell.value, str) and '\n' in cell.value:
                    cell.alignment = Alignment(wrap_text=True)

        wb.save(file_path)
    else:
        raise ValueError(f"不支持的文件格式：{suffix}")
    

        


if __name__ == "__main__":
    file_path = "/home/dev2/PyProject/wangdehua/data/智能客服测试结果.xlsx"
    output_path = "/home/dev2/PyProject/wangdehua/data/test.xlsx"
    write_md_to_excel(output_path, read_excel_to_md(file_path))